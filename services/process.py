from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
import pandas as pd
import numpy as np
import threading

from services.drive_manager import GoogleService
from services.mongodb import DBMongo

class Pipelines:
    def __init__(self, database_service: DBMongo, google_service: GoogleService):
        self.mongo_service = database_service
        self.google_service = google_service

    def getGastos(self):
        table_expenses = self.mongo_service.getGastos()
        table_expenses["Fecha"] = table_expenses["Fecha"].dt.strftime("%Y-%m-%d")
        return table_expenses
    
    def getSummaryByWeek(self):
        sum_expenses = self.mongo_service.getSummaryAmountGastos()
        sum_jornales = self.mongo_service.getSumaryAmountJornales()
        sum_sendMoney = self.mongo_service.getSummaryAmountEnvios()

        df_consolidado = pd.merge(sum_expenses,sum_jornales, on=['Fecha Inicio', 'Fecha Fin'], how='outer')
        df_consolidado = pd.merge(df_consolidado,sum_sendMoney, on=['Fecha Inicio', 'Fecha Fin'], how='outer')
        return df_consolidado
    
    def uploadFile(self, file_info=None):
        data = {"fileDriveId": "", "fileDriveUrl": ""}
        if file_info:
            try:
                tz = timezone(timedelta(hours=-5))
                upload_with_time = datetime.now(tz)
                uploadDate = upload_with_time.strftime("%Y-%m-%d")
                resultado = self.google_service.uploadToDriveByDate(
                    fecha=uploadDate,
                    file_info=file_info
                )
                data["fileDriveId"] = resultado.get("id", "")
                data["fileDriveUrl"] = resultado.get("url", "")
            except Exception as e:
                print(f"Error al subir archivo a Drive: {e}")
        return data

    def postSentMoney(self, data: dict, file_info=None):
        def tarea():
            print(data)
            fecha = data["sentAt"]
            if file_info:
                try:
                    # Subir archivo al Drive y obtener ID y URL
                    resultado = self.google_service.uploadToDriveByDate(
                        fecha=fecha,
                        file_info=file_info
                    )
                    data["fileDriveId"] = resultado["id"]
                    data["fileDriveUrl"] = resultado["url"]

                except Exception as e:
                    print(f"Error al subir archivo a Drive: {e}")
            else:
                data["fileDriveId"] = ""
                data["fileDriveUrl"] = ""
            # Guardar en MongoDB
            self.mongo_service.uploadSendMoney(data)
            print("Envío guardado correctamente")

        threading.Thread(target=tarea, daemon=True).start()

    def getEnvios(self):
        table_sendMoney = self.mongo_service.getEnvios()
        #print(table_sendMoney)
        return table_sendMoney
    
    def getJornales(self):
        table_jornales = self.mongo_service.getJornales()
        #print(table_jornales)
        return table_jornales
    
    def getTransactions(self):
        table_expenses = self.mongo_service.getGastos()
        table_sendMoney = self.mongo_service.getEnvios()
        table_jornales = self.mongo_service.getJornales()
        table_sales = self.mongo_service.getSales()
        
        # Formateando Tabla gastos
        # Filtrar abonos
        df_abono = table_expenses[table_expenses['Producto'] == 'Abono'][['Fecha','Tipo','Producto','Actividad','Descripcion','Monto Total', 'Responsable']].rename(columns={'Producto': 'Nombre', 'Monto Total': 'GastoAbono'})
        # Filtrar los demás gastos
        df_gastos = table_expenses[table_expenses['Producto'] != 'Abono'][['Fecha','Tipo','Producto','Actividad','Descripcion','Monto Total','Responsable']].rename(columns={'Producto': 'Nombre', 'Monto Total': 'Monto'})

        # Formateando Tabla Jornales
        table_jornales = table_jornales[['Fecha Trabajo','Tipo','Trabajador','Actividad', 'Descripcion','Monto Total', 'Responsable', 'Periodo']].rename(columns={
            'Fecha Trabajo' : 'Fecha',
            'Trabajador' : 'Nombre',
            'Monto Total' : 'Jornal'
        })
        table_jornales['JornalMensual'] = np.where(table_jornales['Periodo'] == 'Mensual', table_jornales['Jornal'], None)
        table_jornales['JornalDiario'] = np.where((table_jornales['Periodo'] != 'Mensual') | (table_jornales['Periodo'].isna()), table_jornales['Jornal'], None)

        table_jornales.to_csv("out/jornales_proceed.csv", index=False)

        # Formateando Tabla Enviado
        table_sendMoney = table_sendMoney[['Fecha', 'Tipo','Descripcion', 'Monto Total', 'Responsable']].rename(columns={
            'Descripcion' : 'Nombre',
            'Monto Total' : 'Enviado'
        })
        # Formateando Tabla Gastos
        table_sales = table_sales[['Fecha Venta', 'Tipo', 'Producto', 'Monto', 'Responsable']].rename(columns={
            'Fecha Venta' : 'Fecha',
            'Producto' : 'Nombre',
            'Monto' : 'Venta'
        })

        df_consolidado = pd.concat([df_gastos,df_abono,table_jornales,table_sendMoney,table_sales],axis=0)
        df_consolidado.sort_values(by='Fecha',ascending=True)
        df_consolidado['Actividad'] = df_consolidado['Actividad'].fillna('')
        df_consolidado = df_consolidado[['Fecha', 'Responsable','Tipo', 'Nombre', 'Actividad', 'Descripcion', 'Monto', 'GastoAbono', 'JornalDiario', 'JornalMensual', 'Enviado', 'Venta']]
        df_consolidado.to_csv('out/consolidado2.csv', index=False)
        print(df_consolidado)
        return df_consolidado

    def updateExpenses(self, e_code, data):
        self.mongo_service.update_Expenses(e_code,data)
        return True

    def deleteExpense(self, e_code):
        self.mongo_service.delete_Expense(e_code)
        return True

    def getSales(self):
        table_sales = self.mongo_service.getSales()
        print(table_sales)
        return table_sales
    
    def updateSale(self, v_code, data):
        print(data)
        self.mongo_service.update_Sales(v_code,data)
        return True

    def deleteSale(self, v_code):
        self.mongo_service.delete_Sales(v_code)
        return True
    
    def updateSendMoney(self, s_code, data):
        print(data)
        self.mongo_service.update_SendMoney(s_code,data)
        return True

    def deleteSendMoney(self, s_code):
        self.mongo_service.delete_SendMoney(s_code)
        return True
    
    def updateJornal(self, j_code, data):
        print(data)
        self.mongo_service.update_Jornal(j_code,data)
        return True

    def deleteJornal(self, j_code):
        self.mongo_service.delete_Jornal(j_code)
        return True
    
    def addTotal(self, df_export: pd.DataFrame = None) -> pd.DataFrame:
        total_row = {}

        for col in df_export.columns:
            if pd.api.types.is_numeric_dtype(df_export[col]):
                total_row[col] = df_export[col].sum()
            else:
                total_row[col] = ""

        df_export = pd.concat([df_export, pd.DataFrame([total_row])], ignore_index=True)
        return df_export
    
    def exportSummaryExcelFormatted(self, df_summary: pd.DataFrame, path: str = 'out/export_summary.xlsx', headers: list = None) -> bool:
        static_width = ["Fecha", "Tipo"]
        format_columns_center = headers[1:]

        # Reemplazar titulos 
        df_summary.columns = headers[1:]

        border = Border(
            left=Side(border_style="thin", color="434241"),
            right=Side(border_style="thin", color="434241"),
            top=Side(border_style="thin", color="434241"),
            bottom=Side(border_style="thin", color="434241")
        )
        font_black = Font(color="000000", bold=True)
        font_white = Font(color="FFFFFF", bold=True)

        # Crear un estilo con formato de contabilidad
        format_account_soles = NamedStyle(
            name="accounting_style_soles",
            number_format='"S/" #,##0.00_);[Red]("S/" #,##0.00)'
        )
        format_account_soles.font = Font(name="Calibri", size=11)
        format_account_soles.alignment = Alignment(horizontal="center")

        # Estilos para la fecha
        date_format_style = NamedStyle(
            name="date_format_style",
            number_format="dd/mm/yyyy"  # solo fecha (Perú)
        )
        date_format_style.font = Font(name="Calibri", size=11)
        date_format_style.alignment = Alignment(horizontal="center")

        # Búsqueda de índices
        col_idx_fecha = df_summary.columns.get_loc("Fecha") + 1
        col_idx_tipo = df_summary.columns.get_loc("Tipo") + 1
        col_idx_gasto = df_summary.columns.get_loc("Gasto (S/.)") + 1
        col_idx_gastoabono = df_summary.columns.get_loc("Abono (S/.)") + 1
        col_idx_enviado= df_summary.columns.get_loc("Enviado (S/.)") + 1
        col_idx_venta = df_summary.columns.get_loc("Venta Cacao (S/.)") + 1
        col_idx_jornal = df_summary.columns.get_loc("Jornal (S/.)") + 1
        col_idx_jornal_mensual = df_summary.columns.get_loc("J. Mensual (S/.)") + 1

        # Obteniendo las columnas
        col_letter_fecha = get_column_letter(col_idx_fecha)
        col_letter_tipo = get_column_letter(col_idx_tipo)
        col_letter_gasto = get_column_letter(col_idx_gasto)
        col_letter_gastoabono = get_column_letter(col_idx_gastoabono)
        col_letter_enviado = get_column_letter(col_idx_enviado)
        col_letter_venta = get_column_letter(col_idx_venta)
        col_letter_jornal = get_column_letter(col_idx_jornal)
        col_letter_jornal_mensual = get_column_letter(col_idx_jornal_mensual)

        # Estilos de fill
        fill_1 = PatternFill(start_color="D6D64B", end_color="D6D64B", fill_type="solid")
        fill_2 = PatternFill(start_color="c05d49", end_color="c05d49", fill_type="solid")
        fill_3 = PatternFill(start_color="1F66E0", end_color="1F66E0", fill_type="solid")
        fill_4 = PatternFill(start_color="3cad27", end_color="3cad27", fill_type="solid")
        fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, index=False, sheet_name='Reporte')

            # Accede al workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Reporte']
            worksheet.auto_filter.ref = worksheet.dimensions

            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:  # Primera fila (encabezados)
                cell.fill = header_fill
                cell.font = header_font

            last_row = len(df_summary) + 1  # +1 por el encabezado

            # --- Aplicar bordes y colores en la columna 'Estado' ---
            for row in range(2, last_row + 1):  # Desde la fila 2 hasta la última
                cell_fecha = worksheet[f"{col_letter_fecha}{row}"]
                cell_tipo = worksheet[f"{col_letter_tipo}{row}"]
                cell_gasto = worksheet[f"{col_letter_gasto}{row}"]
                cell_gastoabono = worksheet[f"{col_letter_gastoabono}{row}"]
                cell_enviado = worksheet[f"{col_letter_enviado}{row}"]
                cell_venta = worksheet[f"{col_letter_venta}{row}"]
                cell_jornal = worksheet[f"{col_letter_jornal}{row}"]
                cell_jornal_mensual = worksheet[f"{col_letter_jornal_mensual}{row}"]

                cell_venta.style = format_account_soles
                cell_gasto.style = format_account_soles
                cell_gastoabono.style = format_account_soles
                cell_enviado.style = format_account_soles
                cell_venta.style = format_account_soles
                cell_fecha.style = date_format_style
                cell_jornal.style = format_account_soles
                cell_jornal_mensual.style = format_account_soles

                # Aplicar color en base al tipo
                if cell_tipo.value == "Víveres":
                    cell_tipo.fill = fill_2
                    cell_tipo.font = font_white
                if cell_tipo.value == "Gastos":
                    cell_tipo.fill = fill_2
                    cell_tipo.font = font_white
                if cell_tipo.value == "Jornales": 
                    cell_tipo.fill = fill_3
                    cell_tipo.font = font_white
                if cell_tipo.value == "Efectivo":
                    cell_tipo.fill = fill_1
                    cell_tipo.font = font_black
                if cell_tipo.value == "Venta Cacao":
                    cell_tipo.fill = fill_4
                    cell_tipo.font = font_white

            # --- Ajustar ancho de columnas automáticamente ---
            for col_idx, column in enumerate(df_summary.columns, start=1):
                if column in static_width: 
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = len(column) + 10
                else:
                    max_length = max(
                        df_summary[column].astype(str).map(len).max(),
                        len(column)  # incluye el nombre de la columna
                    )
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = max_length + 10  # padding
                if column in format_columns_center:
                    for row in range(2, len(df_summary) + 2):
                        worksheet.cell(row=row, column=col_idx).alignment = Alignment(horizontal="center")

            # Aplica el borde a todas las celdas de la hoja
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border

            # --- Pintar solo la última columna de gris ---
            last_row = worksheet.max_row
            max_col  = worksheet.max_column

            for c in range(1, max_col + 1):
                worksheet.cell(row=last_row, column=c).fill = fill_gray